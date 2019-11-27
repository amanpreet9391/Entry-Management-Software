import mysql.connector
from openpyxl import Workbook
from entry import timeStamp
from host import host
from visitor import visitors
#from main import connect_to_db
mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  passwd="123456789"				#your mysql Password
)

mycursor = mydb.cursor()

def create_database(i):
    query = "INSERT INTO visitor (Name,email,PhoneNumber,CheckinDate, CheckinTime,CheckoutDate, CheckoutTime) VALUES (%s,%s,%s,%s,%s,%s,%s)"
    values = (i[0],i[1],i[2],i[3][0],i[3][1],i[4][0],i[4][1])
    mycursor.execute(query,values)
    mydb.commit()

mycursor.execute("CREATE DATABASE db1")
mycursor.execute("use db1")
mycursor.execute("CREATE TABLE visitor ( Name TEXT, email VARCHAR(255),PhoneNumber VARCHAR(12) , CheckinDate TEXT ,CheckinTime TEXT, CheckoutDate TEXT, CheckoutTime TEXT)")
V=visitors()
id=input("Enter visitor's email address who wants to checkout : ")
#for i in V:
    #if i[1]==id:
        #print(i)
checkout=timeStamp()
V.append(checkout)

create_database(V)

wb = Workbook()
mycursor.execute("SELECT * FROM visitor")
results = mycursor.fetchall()
ws = wb.create_sheet(0)
#ws.title = visitor
ws.append(mycursor.column_names)
for row in results:
    ws.append(row)
workbook_name = "test_workbook"
wb.save(workbook_name + ".xlsx")

def connect_to_db():
    mydb = mysql.connector.connect(
    host = "localhost",
    user = "root",
    password = "123456789",
    database = "db1"
    )
    return mydb


mydb = connect_to_db()
mycursor = mydb.cursor()

# V=visitors()
# id=input("Enter visitor's email address who wants to checkout : ")
# #for i in V:
#     #if i[1]==id:
#         #print(i)
# checkout=timeStamp()
# V.append(checkout)

# create_datatbase(V)




#wb = Workbook()
# SQL = 'USE ' + db1 + ';'
# mycursor.execute(SQL)
# SQL = 'SELECT * from '+ visitor + ';'
# mycursor.execute(SQL)
# results = mycursor.fetchall()
# ws = wb.create_sheet(0)
# ws.title = visitor
# ws.append(mycursor.column_names)
# for row in results:
#     ws.append(row)

# SQL = 'SELECT * from '+ manufacturer_table_name + ';'
# mycursor.execute(SQL)
# results = mycursor.fetchall()
# ws = wb.create_sheet(0)
# ws.title = manufacturer_table_name
# ws.append(mycursor.column_names)
# for row in results:
#     ws.append(row)

# workbook_name = "test_workbook"
# wb.save(workbook_name + ".xlsx")
